#!/usr/bin/env python3
"""Scan C:\\ top-level folders, calculate sizes, export to Excel."""

from __future__ import annotations

import os
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Hiányzó csomag: pip install openpyxl")

ROOT = Path("C:\\")
OUTPUT_FILE = Path("c_drive_folder_sizes.xlsx")


def human_size(num_bytes: int) -> str:
    size = float(num_bytes)
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if size < 1024 or unit == "TB":
            return f"{size:.2f} {unit}"
        size /= 1024
    return f"{num_bytes} B"


def get_folder_size(path: Path) -> tuple[int, int]:
    """Return (total_bytes, error_count) for a directory tree."""
    total = 0
    errors = 0
    try:
        for entry in os.scandir(path):
            try:
                if entry.is_symlink():
                    continue
                if entry.is_dir(follow_symlinks=False):
                    sub_bytes, sub_errors = get_folder_size(Path(entry.path))
                    total += sub_bytes
                    errors += sub_errors
                else:
                    total += entry.stat(follow_symlinks=False).st_size
            except (PermissionError, OSError):
                errors += 1
    except (PermissionError, OSError):
        errors += 1
    return total, errors


def scan_root(root: Path) -> list[dict]:
    rows: list[dict] = []
    try:
        entries = sorted(root.iterdir(), key=lambda p: p.name.lower())
    except PermissionError:
        sys.exit(f"Nem sikerült olvasni: {root}")

    total_entries = sum(1 for e in entries if e.is_dir(follow_symlinks=False))
    processed = 0

    for entry in entries:
        if not entry.is_dir(follow_symlinks=False):
            continue
        processed += 1
        print(f"[{processed}/{total_entries}] Mérem: {entry.name} ...", end="\r")
        size_bytes, error_count = get_folder_size(entry)
        rows.append(
            {
                "folder": entry.name,
                "path": str(entry),
                "size_bytes": size_bytes,
                "size_human": human_size(size_bytes),
                "errors": error_count,
            }
        )

    print()
    rows.sort(key=lambda r: r["size_bytes"], reverse=True)
    return rows


def write_excel(rows: list[dict], output: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "C meghajtó mappák"

    # --- Colours ---
    header_fill = PatternFill("solid", fgColor="1F4E79")
    alt_fill = PatternFill("solid", fgColor="D6E4F0")
    total_fill = PatternFill("solid", fgColor="2E75B6")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    bold_white = Font(bold=True, color="FFFFFF", size=11)

    headers = ["#", "Mappa neve", "Teljes elérési út", "Méret (olvasható)", "Méret (byte)", "Hibás fájlok"]
    col_widths = [5, 30, 55, 18, 18, 14]

    # Title row
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"C:\\ meghajtó mappaméretek  –  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.font = Font(bold=True, color="FFFFFF", size=13)
    title_cell.fill = PatternFill("solid", fgColor="0D2137")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 22

    # Data rows
    for row_idx, row in enumerate(rows, start=1):
        excel_row = row_idx + 2
        fill = alt_fill if row_idx % 2 == 0 else None
        values = [
            row_idx,
            row["folder"],
            row["path"],
            row["size_human"],
            row["size_bytes"],
            row["errors"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            if fill:
                cell.fill = fill
            if col_idx == 5:  # bytes column → right aligned
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
            elif col_idx in (4, 6):
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    # Total row
    total_row = len(rows) + 3
    total_bytes = sum(r["size_bytes"] for r in rows)
    total_errors = sum(r["errors"] for r in rows)
    total_values = ["", "ÖSSZESEN", "", human_size(total_bytes), total_bytes, total_errors]
    for col_idx, value in enumerate(total_values, start=1):
        cell = ws.cell(row=total_row, column=col_idx, value=value)
        cell.font = bold_white
        cell.fill = total_fill
        if col_idx == 5:
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
        elif col_idx in (4, 6):
            cell.alignment = Alignment(horizontal="right")

    # Freeze panes below header
    ws.freeze_panes = "A3"

    # Auto-filter on header
    ws.auto_filter.ref = f"A2:F{len(rows) + 2}"

    wb.save(output)
    print(f"Excel mentve: {output.resolve()}")


def main() -> None:
    print(f"C:\\ meghajtó mappáinak felmérése...")
    rows = scan_root(ROOT)
    if not rows:
        sys.exit("Nem találtam mappákat.")
    print(f"{len(rows)} mappa megmérve. Excel írása...")
    write_excel(rows, OUTPUT_FILE)
    total = sum(r["size_bytes"] for r in rows)
    print(f"\nTop 5 legnagyobb mappa:")
    for r in rows[:5]:
        print(f"  {r['size_human']:>12}  {r['folder']}")
    print(f"\nÖsszesen: {human_size(total)}")


if __name__ == "__main__":
    main()
